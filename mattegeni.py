# mattegeni.py - Optimerad version med logg i Dokument-mapp
# svårigheten är satt till de krav som barn har i mellanstadiet

import random
import time
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from pathlib import Path
import platform

# --- Konstanter och räknesätt ---
RAKNESATT_DICT = {
    1: "multiplikation",
    2: "division",
    3: "subtraktion",
    4: "addition"
}

INTERVALL = {
    1: (1, 10, 1, 10),
    2: (1, 10, 1, 10),
    3: (1, 50, 1, 50),
    4: (1, 50, 1, 50)
}

LOGG_HEADER = [
    'Användare', 'Veckonummer', 'Månad',
    'Räknesätt', 'Fråga', 'Svar', 'Rätt/Fel', 'Svarstid (s)', 'Rätt (1)/Fel (0)'
]

# --- Funktion: Hämta dokumentmapp beroende på operativsystem ---
def hamta_dokumentmapp():
    system = platform.system()
    if system == "Windows":
        return Path(os.path.join(os.environ["USERPROFILE"], "Documents"))
    elif system == "Darwin":  # macOS
        return Path(os.path.join(Path.home(), "Documents"))
    else:  # Linux
        return Path(os.path.join(Path.home(), "Documents"))

# --- Funktion: Generera 10 frågor ---
def generera_frågor(raknesatt):
    frågor = []
    förekomster = {}
    while len(frågor) < 10:
        if raknesatt == 1:
            a = random.randint(*INTERVALL[1][:2])
            b = random.randint(*INTERVALL[1][2:])
        elif raknesatt == 2:
            b = random.randint(*INTERVALL[2][2:])
            korrekt = random.randint(*INTERVALL[2][:2])
            a = b * korrekt
        elif raknesatt == 3:
            a = random.randint(*INTERVALL[3][:2])
            b = random.randint(*INTERVALL[3][2:])
            if a < b:
                a, b = b, a
        elif raknesatt == 4:
            a = random.randint(*INTERVALL[4][:2])
            b = random.randint(*INTERVALL[4][2:])
        else:
            a = b = 1
        fråga = (a, b)
        nyckel = tuple(sorted(fråga))
        if förekomster.get(nyckel, 0) < 2:
            frågor.append(fråga)
            förekomster[nyckel] = förekomster.get(nyckel, 0) + 1
    return frågor

# --- Funktion: Logga till Excel ---
def logga_rad_xlsx(filnamn, rad):
    if not os.path.exists(filnamn):
        wb = Workbook()
        ws = wb.active
        ws.title = "Logg"
        ws.append(LOGG_HEADER)
        wb.save(filnamn)
    wb = load_workbook(filnamn)
    ws = wb.active
    ws.append(rad)
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2
    wb.save(filnamn)

def ja_eller_nej_fraga(fras):
    while True:
        svar = input(fras).strip().lower()
        if svar == "ja":
            return True
        elif svar == "nej":
            return False
        else:
            print("Nu blev det fel, försök igen. Svara 'ja' eller 'nej'.")

def valj_raknesatt():
    while True:
        print("\nVilket räknesätt vill du träna på idag?")
        for i in range(1, 5):
            print(f"[{i}] - {RAKNESATT_DICT[i]}")
        val = input("Skriv siffran för ditt val: ").strip()
        if val in ["1", "2", "3", "4"]:
            return int(val), RAKNESATT_DICT[int(val)]
        else:
            print("Nu blev det fel, försök igen.")

def skapa_fraga_och_svar(a, b, raknesatt):
    if raknesatt == 1:
        return f"{a} * {b}", a * b
    elif raknesatt == 2:
        return f"{a} / {b}", a // b
    elif raknesatt == 3:
        return f"{a} - {b}", a - b
    elif raknesatt == 4:
        return f"{a} + {b}", a + b

def main():
    print("Välkommen till räkneträningen!")
    namn = input("Vad heter du? ").strip()
    if not namn:
        namn = "mattegeni"

    dokument_path = hamta_dokumentmapp() / "mattegeni_loggar"
    dokument_path.mkdir(exist_ok=True)
    filnamn = dokument_path / f"{namn}.xlsx"

    raknesatt, raknesatt_namn = valj_raknesatt()

    while True:
        frågor = generera_frågor(raknesatt)
        rätt = 0
        fel = 0
        fel_lista = []
        print(f"\nNu kör vi 10 tal med {raknesatt_namn}, {namn}! Lycka till!\n")
        for idx, (a, b) in enumerate(frågor, 1):
            fråga_str, korrekt = skapa_fraga_och_svar(a, b, raknesatt)
            start_tid = time.time()
            while True:
                svar = input(f"{idx}. Vad är {fråga_str}? ")
                svar_tid = time.time() - start_tid
                if svar.lstrip('-').isdigit():
                    svar_int = int(svar)
                    break
                else:
                    print("Det där var ingen siffra, försök igen.")
                    start_tid = time.time()
            nu = datetime.now()
            veckonummer = nu.isocalendar()[1]
            manad = nu.strftime('%B')
            if svar_int == korrekt:
                print("Rätt! Bra jobbat!\n")
                rätt += 1
                rätt_fel = "rätt"
                rf_varde = 1
            else:
                print(f"Fel! Rätt svar är {korrekt}.\n")
                fel += 1
                rätt_fel = "fel"
                rf_varde = 0
                fel_lista.append((fråga_str, svar_int))
            logga_rad_xlsx(
                filnamn,
                [namn, veckonummer, manad, raknesatt_namn, fråga_str, svar_int, rätt_fel, round(svar_tid, 1), rf_varde]
            )
        print(f"\nDen här gången hade du {rätt} rätt och {fel} fel, {namn}.")
        if fel_lista:
            print("Du hade fel på dessa tal:")
            for fråga, svar in fel_lista:
                print(f"  {fråga} (ditt svar: {svar})")
        if rätt >= 7:
            print(f"\nNu kan du vara stolt ditt lilla mattegeni, du har nu tjänat {rätt} poäng!")
        if not ja_eller_nej_fraga("\nVill du köra 10 nya tal? Då får du välja ett nytt räknesätt om du vill (ja/nej): "):
            print(f"Bra kämpat idag! Hej då {namn}!")
            break
        raknesatt, raknesatt_namn = valj_raknesatt()

if __name__ == "__main__":
    main()
